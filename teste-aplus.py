import openpyxl

class MaterialList:
    def __init__(self, material_list_path, aplus_db_path) -> None:
        # Carrega as duas planilhas (lista de materiais e banco Aplus)
        self.lista_materiais_wb = openpyxl.load_workbook(material_list_path)
        self.aplus_db_wb = openpyxl.load_workbook(aplus_db_path)

        # Define a aba ativa (a primeira) de cada planilha
        self.lista_materiais = self.lista_materiais_wb.active
        self.aplus_db = self.aplus_db_wb.active

        # Chama a função que adiciona a coluna de importação
        self.append_new_column()
        
        # Salva as alterações na planilha de materiais
        self.save_list(material_list_path)
        

    def get_aplus_data(self, codigo):
        # Percorre as linhas do banco de dados Aplus (começando da segunda linha)
        for row in self.aplus_db.iter_rows(min_row=2, values_only=True):
            if row[0] == codigo:  # Se o código bate
                peso_unitario = row[2]
                unidade_compra = row[3]
                return peso_unitario, unidade_compra
        # Se não encontrar o material, retorna None
        return None, None

    def calculate_import_quantity(self, qtd, comprimento, peso, peso_unitario, unidade_de_compra):
        # Se a unidade de compra for barra (cada barra tem 6 metros)
        if unidade_de_compra == 'br':
            if comprimento > 0:
                # Faz o cálculo convertendo milímetros para metros e dividindo pelo peso unitário
                return (((comprimento / 1000) / 6) * peso) / peso_unitario
  
        # Se não for barra, faz o cálculo comum
        return (qtd * peso) / peso_unitario
    
    def append_new_column(self):
        # Verifica se a coluna 'QUANTIDADE DE IMPORTAÇÃO' já existe
        if 'QUANTIDADE DE IMPORTAÇÃO' not in [cell.value for cell in self.lista_materiais[1]]:
            # Adiciona o cabeçalho dessa coluna
            self.lista_materiais.cell(row=1, column=self.lista_materiais.max_column + 1).value = 'QUANTIDADE DE IMPORTAÇÃO'

        # Itera sobre as linhas da planilha de materiais
        for row_idx, row in enumerate(self.lista_materiais.iter_rows(min_row=2, values_only=True), start=2):
            # Busca as informações do banco Aplus
            peso_unitario, unidade_compra = self.get_aplus_data(row[0])

            # Se encontrou os dados no banco
            if peso_unitario and unidade_compra:
                # Calcula a quantidade de importação
                quantidade_importacao = self.calculate_import_quantity(row[2], row[3], row[4], peso_unitario, unidade_compra)
                # Adiciona o valor na nova coluna
                self.lista_materiais.cell(row=row_idx, column=self.lista_materiais.max_column).value = quantidade_importacao

        # Print para confirmar que a nova coluna foi adicionada com sucesso
        print("Coluna 'QUANTIDADE DE IMPORTAÇÃO' adicionada com sucesso!")

    def save_list(self, material_list_path):
        # Salva a planilha com as alterações
        self.lista_materiais_wb.save(material_list_path)

        # Print para confirmar que a planilha foi salva com sucesso
        print(f"Planilha '{material_list_path}' salva com sucesso!")

def main() -> None:
    # Define os caminhos para as planilhas
    material_list_path = 'Lista de Materiais.xlsx'
    aplus_db_path = 'Banco de Dados.xlsx'
    
    # Cria o objeto MaterialList e executa as operações
    MaterialList(material_list_path, aplus_db_path)

    # Print para indicar que o processo inteiro foi concluído
    print("Processo finalizado com sucesso!")

if __name__ == '__main__':
    main()
